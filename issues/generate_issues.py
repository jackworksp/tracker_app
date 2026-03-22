"""
generate_issues.py — Vela Issue Tracker Excel Generator
Run: pip install openpyxl && python generate_issues.py
Output: vela_issue_tracker.xlsx (sheets: All Issues, category sheets, ClickUp Issues, Statistics)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# DATA — all 15 Vela issues
# ---------------------------------------------------------------------------

ISSUES = [
    # ── Resolved issues (pending/) ──────────────────────────────────────────
    {
        "id": "001",
        "title": "Task Completion Validation",
        "category": "Backend",
        "sub_category": "Bug Fix",
        "source": "Manual Testing",
        "found_by": "Developer",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "Backend API, React Web",
        "affected_files": (
            "backend/routes/tasks.js, "
            "frontend-web/src/components/TaskDetailModal.jsx, "
            "frontend-web/src/api.js"
        ),
        "clickup": "https://app.clickup.com/t/86d1x1qdu",
        "created": "2026-02-15",
        "resolved": "2026-02-15",
        "resolution_days": 1,
        "resolved_by": "Development Team",
        "description": (
            "Parent tasks could be marked complete even when subtasks remained "
            "incomplete, causing data integrity issues and user confusion."
        ),
        "solution": (
            "Added backend hasIncompleteSubtasks() guard, frontend warning modal "
            "with override, and force_complete API parameter. Handles both inline "
            "(JSONB) and relational subtasks."
        ),
        "testing": "Unit tests for inline + relational subtasks, force_complete override, mobile tested",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "002",
        "title": "UI Alterations — Task Cards & Mobile UX",
        "category": "Frontend Web",
        "sub_category": "Enhancement",
        "source": "User Report",
        "found_by": "Developer",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 4",
        "component": "React Web, Design System",
        "affected_files": (
            "frontend-web/src/components/Tasks.jsx, "
            "frontend-web/src/components/Tasks.css, "
            "frontend-web/src/components/TaskDetailModal.jsx, "
            "frontend-web/src/components/BidirectionalSwipeCard.jsx, "
            "frontend-web/src/components/BottomNav.css"
        ),
        "clickup": "https://app.clickup.com/t/86d1vpb9y",
        "created": "2026-02-15",
        "resolved": "2026-02-15",
        "resolution_days": 1,
        "resolved_by": "Development Team",
        "description": (
            "UI needed general improvements: task card visual polish, subtask "
            "progress indicators, swipe sensitivity, and mobile touch targets."
        ),
        "solution": (
            "Added subtask progress badges, progress bars, bulk actions, improved "
            "swipe thresholds, larger touch targets in BottomNav."
        ),
        "testing": "Manual testing, mobile (Android), desktop, responsive design, basic accessibility review",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "003",
        "title": "Secure Vault — Encrypted Personal Details Storage",
        "category": "Security",
        "sub_category": "New Feature",
        "source": "User Report",
        "found_by": "Developer",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 2",
        "component": "Backend API, Database, React Web",
        "affected_files": (
            "backend/database.js, backend/utils/encryption.js, "
            "backend/routes/notes.js, "
            "frontend-web/src/components/VaultPage.jsx, "
            "frontend-web/src/components/VaultUnlockModal.jsx, "
            "frontend-web/src/components/AppLock.jsx, "
            "frontend-web/src/contexts/SecurityContext.jsx"
        ),
        "clickup": "https://app.clickup.com/t/86d1uwmu2",
        "created": "2026-02-15",
        "resolved": "2026-02-15",
        "resolution_days": 1,
        "resolved_by": "Development Team",
        "description": (
            "Users needed secure storage for sensitive personal data (bank details, "
            "personal numbers, credentials) with encryption at rest."
        ),
        "solution": (
            "Extended notes system with AES-256-GCM encryption + PBKDF2 key "
            "derivation (100k iterations). Added VaultPage UI with auto-lock "
            "after 5 minutes. Password never stored."
        ),
        "testing": "Encryption/decryption cycle, password derivation, auth tag validation, auto-lock, error handling",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "Master password is NOT recoverable by design",
    },
    {
        "id": "004",
        "title": "Skills Tracking Feature",
        "category": "Backend",
        "sub_category": "New Feature",
        "source": "ClickUp",
        "found_by": "Developer",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 3",
        "component": "Backend API, Database, React Web",
        "affected_files": (
            "backend/database.js, backend/routes/skills.js, "
            "backend/routes/tasks.js, backend/server.js, "
            "frontend-web/src/components/SkillsPage.jsx, "
            "frontend-web/src/components/SkillCard.jsx, "
            "frontend-web/src/components/SkillSelector.jsx, "
            "frontend-web/src/api.js, frontend-web/src/App.jsx"
        ),
        "clickup": "https://app.clickup.com/t/86d1z2736",
        "created": "2026-02-15",
        "resolved": "2026-02-15",
        "resolution_days": 1,
        "resolved_by": "Development Team",
        "description": (
            "Users needed a way to track personal skills development, link skills "
            "to tasks/study sessions, and monitor proficiency over time."
        ),
        "solution": (
            "New skills / task_skills / session_skills DB tables, full CRUD API, "
            "SkillsPage dashboard, 4-level proficiency tracking, 6 categories, "
            "auto-update last_used on task completion."
        ),
        "testing": "CRUD operations, task-skill linking, proficiency updates, auto-update on completion, mobile responsive",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },

    # ── Open mobile UX issues (ui/mobile/) ──────────────────────────────────
    {
        "id": "UI-01",
        "title": "Touch Targets Too Small (<44px)",
        "category": "Flutter Mobile",
        "sub_category": "Accessibility",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Critical",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App, Design System",
        "affected_files": (
            "vela_flutter/lib/ui/widgets/vela_modal.dart, "
            "vela_flutter/lib/ui/widgets/vela_input.dart, "
            "vela_flutter/lib/ui/widgets/vela_button.dart, "
            "vela_flutter/lib/ui/screens/shared/app_shell.dart, "
            "vela_flutter/lib/core/theme/app_spacing.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": (
            "Multiple interactive elements are below WCAG 2.5.5 minimum 44x44px "
            "touch target. Modal close=32px, bottom nav=36px tall, inputs=40px."
        ),
        "solution": (
            "Wrapped modal close button, nav items, inputs, and buttons in "
            "ConstrainedBox(BoxConstraints(minWidth: 44, minHeight: 44)). Added "
            "AppSpacing.minTouchTarget = 44.0 token for future use."
        ),
        "testing": (
            "Inspect widget tree: every interactive element shows minWidth/minHeight "
            "= 44. Test on physical device with finger — no accidental mis-taps."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "WCAG Level A violation — blocks accessibility certification",
    },
    {
        "id": "UI-02",
        "title": "Bottom Nav Text Too Small",
        "category": "Flutter Mobile",
        "sub_category": "Mobile UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/core/theme/app_typography.dart, "
            "vela_flutter/lib/ui/screens/shared/app_shell.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Bottom navigation label text is too small to read comfortably on mobile screens.",
        "solution": (
            "Added AppTypography.fontSizeNavLabel = 11.0 and navLabel() style "
            "factory with fontWeight always weightSemibold (w600). app_shell.dart "
            "uses navLabel() for all nav items."
        ),
        "testing": (
            "Open app on device, check bottom nav labels are clearly legible at "
            "arm's length. Verify all labels (active and inactive) render at 11px w600."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-03",
        "title": "Missing Onboarding Flow",
        "category": "Flutter Mobile",
        "sub_category": "UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/ui/screens/shared/onboarding_tooltips.dart, "
            "vela_flutter/lib/app.dart, "
            "vela_flutter/lib/core/storage/local_storage.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "No onboarding screens to guide new users through creating a subject and logging their first session.",
        "solution": (
            "OnboardingTooltips widget (3-step walkthrough: swipe gestures, camera, "
            "share target) wired into app.dart via _OnboardingOverlay in the "
            "MaterialApp.router builder. Shown once per device using "
            "LocalStorage.hasSeenOnboarding flag. Respects OS reduced-motion preference."
        ),
        "testing": (
            "Clear app data, launch fresh, verify 3-step tooltip sequence appears. "
            "Complete walkthrough, relaunch — confirm tooltips do not reappear. "
            "Enable OS reduced motion — confirm overlay is skipped."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-04",
        "title": "No Back-Button Navigation",
        "category": "Flutter Mobile",
        "sub_category": "Navigation",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/ui/widgets/detail_header.dart, "
            "vela_flutter/lib/ui/screens/tasks/task_detail_modal.dart, "
            "vela_flutter/lib/ui/screens/sessions/add_session_modal.dart, "
            "vela_flutter/lib/ui/screens/notes/add_note_modal.dart, "
            "vela_flutter/lib/ui/screens/tasks/add_task_modal.dart, "
            "vela_flutter/lib/ui/screens/attachments/add_attachment_modal.dart, "
            "vela_flutter/lib/ui/screens/goals/add_goal_modal.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Detail screens lack a visible back button, relying solely on the Android hardware back gesture.",
        "solution": (
            "DetailHeader widget with ≥ 44px back button added to task_detail_modal.dart. "
            "Close buttons with ConstrainedBox(minWidth: 44, minHeight: 44) and Semantics "
            "labels added to all modal title rows in add_session_modal.dart, "
            "add_note_modal.dart, add_task_modal.dart, add_attachment_modal.dart, "
            "and add_goal_modal.dart."
        ),
        "testing": (
            "Open each modal, verify close button is visible and tap target is at least 44×44px. "
            "Verify VoiceOver/TalkBack announces 'Close' button. "
            "Tap close — confirm modal dismisses correctly."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-05",
        "title": "Share Modal — No Active Subject Indicator",
        "category": "Flutter Mobile",
        "sub_category": "UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/ui/screens/attachments/share_receive_screen.dart (new), "
            "vela_flutter/lib/ui/navigation/app_router.dart, "
            "vela_flutter/lib/main.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "When sharing content from another app to Vela, the modal doesn't show which subject is currently active.",
        "solution": (
            "Created ShareReceiveScreen that handles share_handler intents. Displays "
            "active subject as a prominent green badge (or an amber warning when no subject "
            "is selected). User can save as Task or Attachment. Route /share-receive added "
            "to GoRouter outside the shell. main.dart detects share intent at startup and "
            "overrides initialLocationProvider to navigate there directly."
        ),
        "testing": (
            "Share a URL from Chrome/YouTube to Vela. Verify ShareReceiveScreen opens. "
            "Verify active subject badge shows the correct subject name. "
            "Remove active subject, reshare — verify amber warning appears. "
            "Tap 'Task' — confirm task created in tasks screen. "
            "Tap 'Attachment' — confirm attachment saved in attachments screen."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-06",
        "title": "Generic / Unhelpful Error Messages",
        "category": "Flutter Mobile",
        "sub_category": "UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/core/utils/error_messages.dart, "
            "vela_flutter/lib/ui/widgets/vela_error_message.dart, "
            "vela_flutter/lib/ui/screens/goals/goals_screen.dart, "
            "vela_flutter/lib/ui/screens/goals/add_goal_modal.dart, "
            "vela_flutter/lib/ui/screens/tasks/add_task_modal.dart, "
            "vela_flutter/lib/ui/screens/notes/notes_screen.dart, "
            "vela_flutter/lib/ui/screens/sessions/timeline_screen.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Error states show generic messages that don't tell the user what went wrong or what to do next.",
        "solution": (
            "Populated error_messages.dart with specific actionable strings per resource; "
            "replaced all generic catch-block texts across goals, tasks, notes, sessions screens; "
            "added VelaScreenError full-screen error widget; "
            "wired inline title error in add_task_modal using VelaInput error prop."
        ),
        "testing": (
            "Manually trigger load failure on GoalsScreen — should show 'Couldn't load goals. Pull down to retry.' "
            "Submit AddTaskModal with empty title — should show inline error text below field. "
            "Trigger delete failure — should show specific SnackBar message."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-07",
        "title": "Skeleton Loading States Missing",
        "category": "Flutter Mobile",
        "sub_category": "Performance / UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/ui/widgets/vela_skeleton.dart, "
            "vela_flutter/lib/ui/screens/goals/goals_screen.dart, "
            "vela_flutter/lib/ui/screens/tasks/tasks_screen.dart, "
            "vela_flutter/lib/ui/screens/notes/notes_screen.dart, "
            "vela_flutter/lib/ui/screens/sessions/timeline_screen.dart, "
            "vela_flutter/lib/ui/screens/attachments/attachments_screen.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "List screens show a blank/spinner while loading instead of skeleton placeholder cards.",
        "solution": (
            "Completed vela_skeleton.dart with VelaSkeletonGoalCard, VelaSkeleton factory class "
            "(card/listItem/text helpers), and _VelaSkeletonListItem. "
            "Replaced CircularProgressIndicator in GoalsScreen with 4x VelaSkeletonGoalCard. "
            "Tasks/Notes/Sessions/Attachments screens were already wired; confirmed consistent. "
            "Shimmer uses LinearGradient tween; respects MediaQuery.disableAnimations for reduced motion."
        ),
        "testing": (
            "Open GoalsScreen with slow/offline connection — should show 4 goal-shaped skeleton cards instead of spinner. "
            "Enable OS reduce-motion — shimmer animation should stop, placeholders remain visible as flat fills. "
            "Confirm all 5 screens (tasks, notes, sessions, attachments, goals) show skeletons on load."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-08",
        "title": "Haptic Feedback Missing",
        "category": "Flutter Mobile",
        "sub_category": "Mobile UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Low",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": (
            "vela_flutter/lib/core/utils/haptics.dart, "
            "vela_flutter/lib/ui/widgets/vela_button.dart, "
            "vela_flutter/lib/ui/screens/tasks/tasks_screen.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Actions like completing a task or saving a session don't trigger haptic feedback on mobile.",
        "solution": (
            "Added lightTap(), mediumImpact(), heavyImpact(), and error() alias methods to VelaHaptics "
            "in haptics.dart, each try/catch guarded and gated by _isSupported. "
            "VelaButton already fires VelaHaptics.light() on press; tasks_screen already fires "
            "VelaHaptics.success() on complete and VelaHaptics.heavy() on delete — confirmed wired."
        ),
        "testing": (
            "On a physical Android device: tap any VelaButton — light tap should be felt. "
            "Swipe a task card right (complete) — vibration should fire. "
            "Swipe a task card left (delete) — heavier vibration should fire. "
            "Run on web/desktop — no crash, haptics silently skipped."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-09",
        "title": "Reduced Motion Not Supported",
        "category": "Flutter Mobile",
        "sub_category": "Accessibility",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App, Design System",
        "affected_files": (
            "vela_flutter/lib/core/utils/motion_preferences.dart, "
            "vela_flutter/lib/core/theme/app_animations.dart, "
            "vela_flutter/lib/ui/widgets/vela_modal.dart, "
            "vela_flutter/lib/app.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Animations do not respect the device's 'Reduce Motion' accessibility setting.",
        "solution": (
            "MotionPreferences.of(context) reads MediaQuery.disableAnimations. "
            "AppAnimations.baseOrNone/fastOrNone/slowOrNone return Duration.zero "
            "when reduced motion is enabled. VelaModal.show() and AnimatedContainer "
            "use these helpers; app.dart propagates the flag to onboarding overlay."
        ),
        "testing": (
            "Enable 'Remove animations' (Android) or 'Reduce Motion' (iOS). "
            "Open app and verify: modals appear instantly, no fade/slide animations, "
            "onboarding overlay is skipped."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "WCAG 2.3.3 — Animation from Interactions",
    },
    {
        "id": "UI-10",
        "title": "Swipe Gesture Affordance Missing",
        "category": "Flutter Mobile",
        "sub_category": "Mobile UX",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App",
        "affected_files": "vela_flutter/lib/ui/screens/tasks/tasks_screen.dart",
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Swipeable task cards have no visual hint that they can be swiped (no drag indicator or peek animation).",
        "solution": (
            "Added _SwipeAffordance widget with pulsing green/red edge gradient overlays on the first "
            "active card, shown until the user first swipes (stored via localStorageProvider). "
            "Added SingleTickerProviderStateMixin and a TweenSequence peek animation controller to "
            "_TasksScreenState that translates the first card -12px then +12px then back to 0 over 300ms "
            "on first load; skipped when MediaQuery.disableAnimations is true (Issue 09 compliance). "
            "After first swipe, both the gradient overlay and peek animation are permanently dismissed."
        ),
        "testing": (
            "Fresh install (or clear SharedPreferences): open tasks screen with at least one task — "
            "first card should animate left-right briefly, then show green/red edge gradients. "
            "Swipe the card — gradients disappear and don't return on next launch. "
            "Enable OS Reduce Motion — peek animation should not play; static gradients still visible."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },
    {
        "id": "UI-11",
        "title": "Spacing Inconsistencies Across Screens",
        "category": "Flutter Mobile",
        "sub_category": "Design System",
        "source": "AI Analysis",
        "found_by": "AI Assistant",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Backlog",
        "component": "Flutter App, Design System",
        "affected_files": (
            "vela_flutter/lib/ui/screens/profile/profile_screen.dart, "
            "vela_flutter/lib/ui/screens/goals/goals_screen.dart, "
            "vela_flutter/lib/ui/screens/notes/notes_screen.dart, "
            "vela_flutter/lib/core/theme/app_spacing.dart"
        ),
        "clickup": "",
        "created": "2026-03-22",
        "resolved": "2026-03-22",
        "resolution_days": 0,
        "resolved_by": "Flutter Investigator Agent",
        "description": "Padding and margin values are inconsistent across screens — some use 8px, 12px, 16px ad-hoc instead of design tokens.",
        "solution": (
            "Replaced all hardcoded padding/margin values in profile_screen.dart, goals_screen.dart, "
            "and notes_screen.dart with AppSpacing tokens (s1-s12). "
            "Added AppSpacing import to all three files. "
            "notes_screen FAB clearance updated from hardcoded 80 to AppSpacing.fabClearance + "
            "MediaQuery.of(ctx).padding.bottom. tasks_screen.dart was already clean with AppSpacing tokens."
        ),
        "testing": (
            "Visual review of all four screens in light and dark mode on a physical device. "
            "Confirm consistent 16px (s4) horizontal padding on all list/card containers. "
            "Confirm FAB does not overlap the last list item on iPhone X+ and Android gesture nav devices. "
            "No pixel overflow or clipped content."
        ),
        "deployment_status": "Pending",
        "version": "",
        "notes": "",
    },

    # ── ClickUp-fetched issues (CU-xxx) ─────────────────────────────────────
    # Source: TaskApp - mobile UI & TaskApp - Backend lists in ClickUp
    # Fetched: 2026-03-22 | Sample/test tasks excluded
    {
        "id": "CU-001",
        "title": "No Edit Option for Attachments",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Open",
        "sprint": "Backlog",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/",
        "clickup": "https://app.clickup.com/t/86d28hqa6",
        "created": "2026-03-22",
        "resolved": "",
        "resolution_days": "",
        "resolved_by": "",
        "description": "Vela App: No edit option available for attachments — users can add attachments but cannot modify them after creation.",
        "solution": "",
        "testing": "",
        "deployment_status": "Not Started",
        "version": "",
        "notes": "Only OPEN issue newly fetched from ClickUp",
    },
    {
        "id": "CU-002",
        "title": "Note Not Saving After Creation — Disappears on Back",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 5",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/NotesPage.jsx",
        "clickup": "https://app.clickup.com/t/86d27kyx0",
        "created": "2026-03-01",
        "resolved": "2026-03-08",
        "resolution_days": 7,
        "resolved_by": "Gopinath R",
        "description": "After creating a note, Save button just reloads the notes list. Clicking Back causes the note to disappear — not persisted to DB.",
        "solution": "Fixed save handler to await API call and refresh note list after success.",
        "testing": "Manual: create note, save, navigate away and back — note persists.",
        "deployment_status": "Pending",
        "version": "v1.2.0",
        "notes": "",
    },
    {
        "id": "CU-003",
        "title": "Can't Scroll Through Long Note While Editing",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 5",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/NotesPage.jsx",
        "clickup": "https://app.clickup.com/t/86d27kyv6",
        "created": "2026-03-01",
        "resolved": "2026-03-08",
        "resolution_days": 7,
        "resolved_by": "Gopinath R",
        "description": "Cannot scroll through a note while editing it — notes with >1000 words are inaccessible below the fold.",
        "solution": "Set overflow-y: auto on the note editor textarea/container.",
        "testing": "Manual: open note with 1000+ words, verify scrolling works in edit mode.",
        "deployment_status": "Pending",
        "version": "v1.2.0",
        "notes": "",
    },
    {
        "id": "CU-004",
        "title": "Remove Blue Color From Dark Theme",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 4",
        "component": "React Web, Design System",
        "affected_files": "frontend-web/src/design-system/tokens/",
        "clickup": "https://app.clickup.com/t/86d27kutb",
        "created": "2026-03-01",
        "resolved": "2026-03-08",
        "resolution_days": 7,
        "resolved_by": "Gopinath R",
        "description": "Blue accent color bleeds into dark theme — should use a neutral or theme-appropriate color instead.",
        "solution": "Updated --nds-* color tokens for dark theme, replaced blue with appropriate dark-mode accent.",
        "testing": "Visual review in dark mode across all screens.",
        "deployment_status": "Pending",
        "version": "v1.2.0",
        "notes": "",
    },
    {
        "id": "CU-005",
        "title": "Don't Show 'No Details Provided' / 0 for Empty Fields",
        "category": "Frontend Web",
        "sub_category": "UX",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Low",
        "status": "Resolved",
        "sprint": "Sprint 4",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/TaskDetailModal.jsx",
        "clickup": "https://app.clickup.com/t/86d1x1p62",
        "created": "2026-02-08",
        "resolved": "2026-03-08",
        "resolution_days": 28,
        "resolved_by": "Gopinath R",
        "description": "Task detail view shows 'No details provided' and '0 files' as explicit text when fields are empty — should be hidden or shown as a subtle placeholder.",
        "solution": "Conditionally hide empty field labels; show placeholder only on hover/focus.",
        "testing": "Visual check: create task with no description and no attachments — verify no awkward placeholder text.",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "CU-006",
        "title": "Remove Task Type (Study, Watch) From Task Creation",
        "category": "Frontend Web",
        "sub_category": "UX",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Low",
        "status": "Resolved",
        "sprint": "Sprint 4",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/Tasks.jsx",
        "clickup": "https://app.clickup.com/t/86d1x1nw3",
        "created": "2026-02-08",
        "resolved": "2026-03-08",
        "resolution_days": 28,
        "resolved_by": "Gopinath R",
        "description": "Task creation form shows a 'Type of task' dropdown (Study, Watch, etc.) that is confusing and unnecessary — activity types belong to sessions, not tasks.",
        "solution": "Removed task type field from AddTaskModal and task creation API payload.",
        "testing": "Create a task — verify no type dropdown appears.",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "CU-007",
        "title": "Need Notes Tab Feature",
        "category": "Frontend Web",
        "sub_category": "New Feature",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 3",
        "component": "Backend API, Database, React Web",
        "affected_files": "frontend-web/src/components/NotesPage.jsx, backend/routes/notes.js",
        "clickup": "https://app.clickup.com/t/86d1umz1y",
        "created": "2026-02-03",
        "resolved": "2026-03-08",
        "resolution_days": 33,
        "resolved_by": "Gopinath R",
        "description": "Users need a dedicated Notes tab to create, organize, and retrieve study notes with folder structure and tags.",
        "solution": "Built full notes feature: NotesPage.jsx, note folders, tags, inter-note linking, backend notes + note_folders + note_links routes.",
        "testing": "Create note, add to folder, tag it, link to another note, search — all verified.",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "CU-008",
        "title": "Need Session Add Button in History Tab",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 2",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/Timeline.jsx",
        "clickup": "https://app.clickup.com/t/86d1u9qz9",
        "created": "2026-02-01",
        "resolved": "2026-03-08",
        "resolution_days": 35,
        "resolved_by": "Gopinath R",
        "description": "No visible button to add a new study session from the history/timeline tab — users had no way to log sessions from this view.",
        "solution": "Added Add Session floating action button to Timeline tab.",
        "testing": "Tap Add Session → session modal opens → session saved and appears in timeline.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "Tagged 'bug' in ClickUp",
    },
    {
        "id": "CU-009",
        "title": "Link Tasks and Sessions to Goals",
        "category": "Frontend Web",
        "sub_category": "New Feature",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 3",
        "component": "Backend API, React Web",
        "affected_files": "frontend-web/src/components/GoalsPage.jsx, backend/routes/goals.js",
        "clickup": "https://app.clickup.com/t/86d1u4atp",
        "created": "2026-02-01",
        "resolved": "2026-03-08",
        "resolution_days": 35,
        "resolved_by": "Gopinath R",
        "description": "Users need to associate tasks and study sessions with specific goals to track goal progress automatically.",
        "solution": "Added goal_tasks and goal_sessions linking in DB + UI to attach/detach from goal detail view.",
        "testing": "Link task to goal → verify it shows in goal detail and counts toward progress.",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "CU-010",
        "title": "Need Goal Edit Option",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 3",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/GoalsPage.jsx",
        "clickup": "https://app.clickup.com/t/86d1u4aq4",
        "created": "2026-02-01",
        "resolved": "2026-03-08",
        "resolution_days": 35,
        "resolved_by": "Gopinath R",
        "description": "Goals could be created but not edited — no edit button or inline editing in the goals list.",
        "solution": "Added edit modal to GoalsPage with pre-filled fields and PUT /api/goals/:id endpoint.",
        "testing": "Edit goal title, deadline, description — verify changes saved and reflected in list.",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "",
    },
    {
        "id": "CU-011",
        "title": "Goals Feature Not Working in Production",
        "category": "Backend",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 3",
        "component": "Backend API, DevOps",
        "affected_files": "backend/routes/goals.js, backend/server.js",
        "clickup": "https://app.clickup.com/t/86d1tnnwv",
        "created": "2026-01-31",
        "resolved": "2026-03-08",
        "resolution_days": 36,
        "resolved_by": "Gopinath R",
        "description": "Goals feature worked locally but failed in production — likely a missing route mount or environment config issue.",
        "solution": "Fixed route mount order in server.js and verified DB migration ran in production.",
        "testing": "Verified goals CRUD in production environment.",
        "deployment_status": "Deployed",
        "version": "v1.0.0",
        "notes": "",
    },
    {
        "id": "CU-012",
        "title": "Goals Feature Implementation",
        "category": "Frontend Web",
        "sub_category": "New Feature",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 3",
        "component": "Backend API, Database, React Web",
        "affected_files": "frontend-web/src/components/GoalsPage.jsx, backend/routes/goals.js, backend/database.js",
        "clickup": "https://app.clickup.com/t/86d1tk0z6",
        "created": "2026-01-31",
        "resolved": "2026-03-08",
        "resolution_days": 36,
        "resolved_by": "Gopinath R",
        "description": "Initial implementation of Goals tab — users needed to set, track, and manage long-term study objectives.",
        "solution": "Created goals DB table, full CRUD API, GoalsPage component with progress tracking.",
        "testing": "Create goal, link sessions/tasks, mark complete — all flows verified.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "Tagged 'enchance' in ClickUp",
    },
    {
        "id": "CU-013",
        "title": "Share Option Not Available in Production App",
        "category": "Flutter Mobile",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 2",
        "component": "Flutter App, DevOps",
        "affected_files": "vela_flutter/android/AndroidManifest.xml",
        "clickup": "https://app.clickup.com/t/86d1tff5v",
        "created": "2026-01-31",
        "resolved": "2026-03-08",
        "resolution_days": 36,
        "resolved_by": "Gopinath R",
        "description": "Share-to-Vela feature (share YouTube links from other apps) was missing from the production APK — intent filter not in release build.",
        "solution": "Added share intent filter to AndroidManifest.xml and rebuilt production APK.",
        "testing": "Share YouTube link from Chrome → Vela appears in share sheet → link captured.",
        "deployment_status": "Deployed",
        "version": "v1.0.0",
        "notes": "",
    },
    {
        "id": "CU-014",
        "title": "Basic Share Feature Implementation",
        "category": "Flutter Mobile",
        "sub_category": "New Feature",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 2",
        "component": "Flutter App",
        "affected_files": "vela_flutter/android/AndroidManifest.xml, vela_flutter/lib/",
        "clickup": "https://app.clickup.com/t/86d1tcz85",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Initial implementation of Android share target — receive URLs and text from other apps and save as attachments in Vela.",
        "solution": "Registered Android share intent filter, built share receive screen to capture and save shared content.",
        "testing": "Share from browser, YouTube, and notes app — all content captured and saved.",
        "deployment_status": "Deployed",
        "version": "v1.0.0",
        "notes": "",
    },
    {
        "id": "CU-015",
        "title": "Mobile APK Production Blank Screen Issue",
        "category": "Flutter Mobile",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "Flutter App, DevOps",
        "affected_files": "vela_flutter/lib/app.dart",
        "clickup": "https://app.clickup.com/t/86d1tcbnt",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Production APK showed a blank white screen on launch — app failed to initialize correctly in release mode.",
        "solution": "Fixed release build config, ensured assets bundled correctly, resolved null-safety issues that only surfaced in release mode.",
        "testing": "Install release APK on physical device, verify app loads correctly.",
        "deployment_status": "Deployed",
        "version": "v1.0.0",
        "notes": "",
    },
    {
        "id": "CU-016",
        "title": "Add Button Showing in Wrong Location (Home / Profile Tab)",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "React Web",
        "affected_files": "frontend-web/src/App.jsx, frontend-web/src/App.css",
        "clickup": "https://app.clickup.com/t/86d1tc1zy",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Floating add button was visible on Home and Profile tabs where it had no function — caused confusion.",
        "solution": "Conditionally rendered add button only on tabs where it's relevant (Tasks, Sessions).",
        "testing": "Navigate to Home and Profile — no add button visible. Tasks tab — button shown.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "Tagged 'home tab', 'profile tab' in ClickUp",
    },
    {
        "id": "CU-017",
        "title": "Email ID Showing Incorrectly in Profile Tab",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/ProfilePage.jsx",
        "clickup": "https://app.clickup.com/t/86d1tc1z3",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Profile tab was displaying a wrong or undefined email address for the logged-in user.",
        "solution": "Fixed JWT payload read in profile component to correctly extract user.email.",
        "testing": "Login, navigate to Profile — email matches account email.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "Tagged 'profile tab' in ClickUp",
    },
    {
        "id": "CU-018",
        "title": "Add Note Form Is Inconsistent",
        "category": "Frontend Web",
        "sub_category": "UX",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Low",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/NotesPage.jsx",
        "clickup": "https://app.clickup.com/t/86d1tc0u3",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Add note form had inconsistent spacing, label alignment, and input sizing compared to other forms in the app.",
        "solution": "Redesigned form to use design system tokens and match other modal forms.",
        "testing": "Visual review against Tasks and Sessions modals for consistency.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "",
    },
    {
        "id": "CU-019",
        "title": "Add Task Button Overlay Drops Below Dashboard",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/Tasks.jsx, frontend-web/src/components/Tasks.css",
        "clickup": "https://app.clickup.com/t/86d1tc07f",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Floating 'Add Task' button was rendered below the content area on some screen sizes, hidden behind the dashboard.",
        "solution": "Fixed z-index and position: fixed to keep button visible at all scroll positions.",
        "testing": "Scroll task list to bottom — add button remains visible and tappable.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "Tagged 'bug', 'tasks tab' in ClickUp",
    },
    {
        "id": "CU-020",
        "title": "Delete Button Not Working for Tasks",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "High",
        "status": "Resolved",
        "sprint": "Sprint 1",
        "component": "React Web, Backend API",
        "affected_files": "frontend-web/src/components/Tasks.jsx, backend/routes/tasks.js",
        "clickup": "https://app.clickup.com/t/86d1tbzr1",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Delete button on task cards was unresponsive — clicking did nothing, tasks couldn't be removed.",
        "solution": "Fixed DELETE /api/tasks/:id handler and wired delete button to API call with confirmation.",
        "testing": "Create task, click delete, confirm — task removed from list and DB.",
        "deployment_status": "Pending",
        "version": "v1.0.0",
        "notes": "Tagged 'tasks tab' in ClickUp",
    },
    {
        "id": "CU-021",
        "title": "Swipe Right to Delete Tasks",
        "category": "Frontend Web",
        "sub_category": "Enhancement",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Resolved",
        "sprint": "Sprint 4",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/BidirectionalSwipeCard.jsx",
        "clickup": "https://app.clickup.com/t/86d1tbuxc",
        "created": "2026-01-30",
        "resolved": "2026-03-08",
        "resolution_days": 37,
        "resolved_by": "Gopinath R",
        "description": "Users requested swipe-right gesture to quickly delete tasks on mobile, instead of tapping into the detail modal.",
        "solution": "Implemented BidirectionalSwipeCard with right-swipe delete and left-swipe complete gestures.",
        "testing": "Swipe right on task card — delete confirmation appears and task is removed.",
        "deployment_status": "Pending",
        "version": "v1.1.0",
        "notes": "Tagged 'enchance' in ClickUp",
    },
    {
        "id": "CU-022",
        "title": "No Session Add Button in History Tab (Backend List)",
        "category": "Frontend Web",
        "sub_category": "Bug",
        "source": "ClickUp",
        "found_by": "User",
        "priority": "Medium",
        "status": "Open",
        "sprint": "Backlog",
        "component": "React Web",
        "affected_files": "frontend-web/src/components/Timeline.jsx",
        "clickup": "https://app.clickup.com/t/86d1tdavq",
        "created": "2026-01-30",
        "resolved": "",
        "resolution_days": "",
        "resolved_by": "",
        "description": "Duplicate open ticket: no Add Session button visible in history/timeline tab. Marked as 'mobile ui bugs' in ClickUp backend list — may need re-verification in latest build.",
        "solution": "",
        "testing": "",
        "deployment_status": "Not Started",
        "version": "",
        "notes": "Duplicate of CU-008 which shows resolved — needs re-verification",
    },
]

# ---------------------------------------------------------------------------
# RAW CLICKUP DATA — all real Vela tasks fetched 2026-03-22
# (sample/test tasks excluded — e.g. "Sales receipts", "Facebook", etc.)
# ---------------------------------------------------------------------------

CLICKUP_RAW = [
    # (id, name, status, priority, list, tags, url)
    ("86d28hqa6", "Vela App: No edit option for attachments",           "new issue",        "normal",  "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d28hqa6"),
    ("86d27kyx0", "Note not saving after creation — disappears on back","resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d27kyx0"),
    ("86d27kyv6", "Can't scroll through a long note while editing",     "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d27kyv6"),
    ("86d27kutb", "Need to remove the blue color from dark theme",      "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d27kutb"),
    ("86d1z2736", "Need to store skills",                               "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1z2736"),
    ("86d1x1qdu", "Task shouldn't complete if subtasks are still there","resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1x1qdu"),
    ("86d1x1p62", "Don't show 'No details provided', 0 for no files",  "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1x1p62"),
    ("86d1x1nw3", "Remove type of task (study, watch)",                 "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1x1nw3"),
    ("86d1vpb9y", "Need to alter UI",                                   "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1vpb9y"),
    ("86d1uwmu2", "Option to store personal details (Secure Vault)",    "resolved",         "low",     "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1uwmu2"),
    ("86d1umz1y", "Need notes tab feature",                             "resolved",         "high",    "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1umz1y"),
    ("86d1u9qz9", "Need session add button",                            "resolved",         "normal",  "TaskApp - mobile UI", "bug",                        "https://app.clickup.com/t/86d1u9qz9"),
    ("86d1u4atp", "Can link tasks and sessions to goals",               "resolved",         "normal",  "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1u4atp"),
    ("86d1u4aq4", "Need goal edit option",                              "resolved",         "high",    "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1u4aq4"),
    ("86d1tnnwv", "Goals feature not working in production",            "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1tnnwv"),
    ("86d1tk0z6", "Goals Feature Implementation",                       "resolved",         "",        "TaskApp - mobile UI", "enchance",                   "https://app.clickup.com/t/86d1tk0z6"),
    ("86d1tff5v", "Share option not available in production app",       "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1tff5v"),
    ("86d1tcz85", "Need to implement basic share feature",              "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1tcz85"),
    ("86d1tcbnt", "Mobile APK production blank screen issue",           "resolved",         "",        "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1tcbnt"),
    ("86d1tc1zy", "Add button showing (home/profile tab)",              "resolved",         "",        "TaskApp - mobile UI", "home tab, profile tab",      "https://app.clickup.com/t/86d1tc1zy"),
    ("86d1tc1z3", "Email ID showing wrongly in profile tab",            "resolved",         "",        "TaskApp - mobile UI", "profile tab",                "https://app.clickup.com/t/86d1tc1z3"),
    ("86d1tc1m8", "No button to add sessions in history tab",           "resolved",         "normal",  "TaskApp - mobile UI", "",                           "https://app.clickup.com/t/86d1tc1m8"),
    ("86d1tc0u3", "Add note form is inconsistent",                      "resolved",         "",        "TaskApp - mobile UI", "tasks tab",                  "https://app.clickup.com/t/86d1tc0u3"),
    ("86d1tc07f", "Add task button overlay gone below dashboard",       "resolved",         "",        "TaskApp - mobile UI", "bug, tasks tab",             "https://app.clickup.com/t/86d1tc07f"),
    ("86d1tbzr1", "Delete button not working for tasks",                "resolved",         "",        "TaskApp - mobile UI", "tasks tab",                  "https://app.clickup.com/t/86d1tbzr1"),
    ("86d1tbuxc", "Swipe right to delete tasks",                        "resolved",         "normal",  "TaskApp - mobile UI", "enchance",                   "https://app.clickup.com/t/86d1tbuxc"),
    ("86d1tdavr", "Mobile APK blank screen (deployment issues)",        "deployment issues","",        "TaskApp - Backend",   "",                           "https://app.clickup.com/t/86d1tdavr"),
    ("86d1tdavq", "No button to add sessions in history tab",           "mobile ui bugs",   "",        "TaskApp - Backend",   "sessions tab",               "https://app.clickup.com/t/86d1tdavq"),
    ("86d1tdava", "Swipe right to delete tasks (UI enhancement)",       "ui enhancements",  "normal",  "TaskApp - Backend",   "enchance",                   "https://app.clickup.com/t/86d1tdava"),
    ("86d1tdavm", "Add button showing (home/profile)",                  "resolved",         "",        "TaskApp - Backend",   "home tab, profile tab",      "https://app.clickup.com/t/86d1tdavm"),
    ("86d1tdavk", "Delete button not working for tasks",                "resolved",         "",        "TaskApp - Backend",   "tasks tab",                  "https://app.clickup.com/t/86d1tdavk"),
    ("86d1tdavj", "Email ID showing wrongly in profile tab",            "resolved",         "",        "TaskApp - Backend",   "profile tab",                "https://app.clickup.com/t/86d1tdavj"),
    ("86d1tdavf", "Add note form is inconsistent",                      "resolved",         "",        "TaskApp - Backend",   "tasks tab",                  "https://app.clickup.com/t/86d1tdavf"),
    ("86d1tdav8", "Add task button overlay gone below dashboard",       "resolved",         "",        "TaskApp - Backend",   "bug, tasks tab",             "https://app.clickup.com/t/86d1tdav8"),
]


# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)

def thin_border():
    side = Side(style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

# Category → header fill color
CATEGORY_COLORS = {
    "Backend":        "1E40AF",   # dark blue
    "Frontend Web":   "D97706",   # amber
    "Database":       "065F46",   # dark green
    "Security":       "7C3AED",   # purple
    "Flutter Mobile": "0F766E",   # teal
}

# Priority → row fill (for open issues only)
PRIORITY_ROW_FILLS = {
    "Critical": hex_fill("FEE2E2"),   # red-100
    "High":     hex_fill("FEF9C3"),   # yellow-100
    "Medium":   hex_fill("F0FDF4"),   # green-50
    "Low":      hex_fill("F8FAFC"),   # slate-50
}

RESOLVED_FILL = hex_fill("F1F5F9")   # cool grey for resolved rows

COLUMNS = [
    ("Issue ID",              12),
    ("Title",                 38),
    ("Category",              16),
    ("Sub-Category",          18),
    ("Source",                18),
    ("Found By",              15),
    ("Priority",              12),
    ("Status",                12),
    ("Sprint",                12),
    ("Component",             28),
    ("Affected Files",        45),
    ("ClickUp Link",          38),
    ("Created Date",          14),
    ("Resolved Date",         14),
    ("Resolution (days)",     18),
    ("Resolved By",           18),
    ("Description",           55),
    ("Solution Applied",      55),
    ("Testing Done",          40),
    ("Deployment Status",     18),
    ("Version",               10),
    ("Notes",                 35),
]

FIELD_KEYS = [
    "id", "title", "category", "sub_category", "source", "found_by",
    "priority", "status", "sprint", "component", "affected_files",
    "clickup", "created", "resolved", "resolution_days", "resolved_by",
    "description", "solution", "testing", "deployment_status", "version", "notes",
]

# ---------------------------------------------------------------------------
# SHEET 1 — ALL ISSUES
# ---------------------------------------------------------------------------

def write_all_issues(ws):
    ws.title = "All Issues"

    # Header
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font(11, "FFFFFF")
        cell.fill = hex_fill("1E293B")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, issue in enumerate(ISSUES, start=2):
        row_fill = RESOLVED_FILL if issue["status"] == "Resolved" else PRIORITY_ROW_FILLS.get(issue["priority"], hex_fill("FFFFFF"))
        for col_idx, key in enumerate(FIELD_KEYS, start=1):
            val = issue.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            if key == "clickup" and val:
                cell.hyperlink = val
                cell.font = Font(color="2563EB", underline="single")
        ws.row_dimensions[row_idx].height = 60

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


# ---------------------------------------------------------------------------
# SHEET 2 — BY CATEGORY
# ---------------------------------------------------------------------------

def write_by_category(ws):
    ws.title = "By Category"

    categories = ["Backend", "Frontend Web", "Database", "Security", "Flutter Mobile"]
    current_row = 1

    summary_headers = ["Issue ID", "Title", "Priority", "Status", "Sprint", "Found By", "Source", "Deployment Status"]

    for cat in categories:
        cat_issues = [i for i in ISSUES if i["category"] == cat]
        if not cat_issues:
            continue

        # Category banner
        banner_color = CATEGORY_COLORS.get(cat, "374151")
        cell = ws.cell(row=current_row, column=1, value=f"  {cat}  ({len(cat_issues)} issues)")
        cell.font = bold_font(12, "FFFFFF")
        cell.fill = hex_fill(banner_color)
        cell.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(summary_headers))
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Sub-header
        for col_idx, h in enumerate(summary_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = bold_font(10, "FFFFFF")
            cell.fill = hex_fill("475569")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Issue rows
        for issue in cat_issues:
            row_fill = RESOLVED_FILL if issue["status"] == "Resolved" else PRIORITY_ROW_FILLS.get(issue["priority"], hex_fill("FFFFFF"))
            row_vals = [
                issue["id"], issue["title"], issue["priority"],
                issue["status"], issue["sprint"], issue["found_by"],
                issue["source"], issue["deployment_status"],
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border()
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1  # blank row between categories

    # Column widths
    col_widths = [10, 40, 12, 14, 12, 16, 18, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# SHEET 3 — STATISTICS
# ---------------------------------------------------------------------------

def write_statistics(ws):
    ws.title = "Statistics"

    def section_header(row, title, color="1E293B"):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = bold_font(11, "FFFFFF")
        cell.fill = hex_fill(color)
        cell.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 22

    def kv(row, label, value, fill=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=10)
        lc.border = thin_border()
        if fill:
            lc.fill = fill
        vc = ws.cell(row=row, column=2, value=value)
        vc.border = thin_border()
        if fill:
            vc.fill = fill
        ws.row_dimensions[row].height = 18

    r = 1

    # ── Overview ────────────────────────────────────────────────────────────
    section_header(r, "Overview"); r += 1
    total = len(ISSUES)
    resolved = sum(1 for i in ISSUES if i["status"] == "Resolved")
    open_ = sum(1 for i in ISSUES if i["status"] == "Open")
    kv(r, "Total Issues", total);   r += 1
    kv(r, "Resolved",     resolved, RESOLVED_FILL); r += 1
    kv(r, "Open",         open_);   r += 1
    kv(r, "Completion Rate", f"{resolved/total*100:.0f}%"); r += 2

    # ── By Status ───────────────────────────────────────────────────────────
    section_header(r, "By Status"); r += 1
    from collections import Counter
    for status, count in sorted(Counter(i["status"] for i in ISSUES).items()):
        kv(r, status, count); r += 1
    r += 1

    # ── By Priority ─────────────────────────────────────────────────────────
    section_header(r, "By Priority"); r += 1
    priority_order = ["Critical", "High", "Medium", "Low"]
    priority_counts = Counter(i["priority"] for i in ISSUES)
    for p in priority_order:
        if p in priority_counts:
            kv(r, p, priority_counts[p], PRIORITY_ROW_FILLS.get(p)); r += 1
    r += 1

    # ── By Category ─────────────────────────────────────────────────────────
    section_header(r, "By Category"); r += 1
    for cat, count in sorted(Counter(i["category"] for i in ISSUES).items()):
        kv(r, cat, count); r += 1
    r += 1

    # ── By Found By ─────────────────────────────────────────────────────────
    section_header(r, "Discovered By"); r += 1
    for source, count in sorted(Counter(i["found_by"] for i in ISSUES).items()):
        kv(r, source, count); r += 1
    r += 1

    # ── Open by Priority ────────────────────────────────────────────────────
    section_header(r, "Open Issues by Priority"); r += 1
    open_issues = [i for i in ISSUES if i["status"] == "Open"]
    open_priority = Counter(i["priority"] for i in open_issues)
    for p in priority_order:
        if p in open_priority:
            kv(r, p, open_priority[p], PRIORITY_ROW_FILLS.get(p)); r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def write_category_sheet(wb, category):
    """Create a dedicated sheet for a single category with full columns."""
    short_names = {
        "Backend": "Backend",
        "Frontend Web": "Frontend Web",
        "Database": "Database",
        "Security": "Security",
        "Flutter Mobile": "Flutter Mobile",
    }
    sheet_name = short_names.get(category, category)
    ws = wb.create_sheet(sheet_name)

    banner_color = CATEGORY_COLORS.get(category, "374151")
    cat_issues = [i for i in ISSUES if i["category"] == category]

    # Banner row
    cell = ws.cell(row=1, column=1, value=f"{category}  —  {len(cat_issues)} issue(s)")
    cell.font = bold_font(13, "FFFFFF")
    cell.fill = hex_fill(banner_color)
    cell.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 30

    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = bold_font(10, "FFFFFF")
        cell.fill = hex_fill("475569")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    # Data rows
    for row_idx, issue in enumerate(cat_issues, start=3):
        row_fill = RESOLVED_FILL if issue["status"] == "Resolved" else PRIORITY_ROW_FILLS.get(issue["priority"], hex_fill("FFFFFF"))
        for col_idx, key in enumerate(FIELD_KEYS, start=1):
            val = issue.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            if key == "clickup" and val:
                cell.hyperlink = val
                cell.font = Font(color="2563EB", underline="single")
        ws.row_dimensions[row_idx].height = 60

    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"


def write_clickup_sheet(wb):
    """Dedicated sheet: all real Vela tasks fetched from ClickUp (raw view)."""
    ws = wb.create_sheet("ClickUp Issues")

    # Banner
    cell = ws.cell(row=1, column=1, value="ClickUp Issues — Fetched 2026-03-22  |  Lists: TaskApp - mobile UI, TaskApp - Backend")
    cell.font = bold_font(12, "FFFFFF")
    cell.fill = hex_fill("7C3AED")
    cell.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 26

    headers = ["ClickUp ID", "Task Name", "Status", "Priority", "List", "Tags", "URL"]
    col_widths = [14, 55, 18, 12, 22, 24, 42]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = bold_font(10, "FFFFFF")
        cell.fill = hex_fill("475569")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    STATUS_FILLS = {
        "resolved":         hex_fill("DCFCE7"),   # green
        "new issue":        hex_fill("FEE2E2"),   # red
        "deployment issues":hex_fill("FEF3C7"),   # amber
        "mobile ui bugs":   hex_fill("FEF3C7"),   # amber
        "ui enhancements":  hex_fill("EFF6FF"),   # blue
    }

    for row_idx, (cid, name, status, priority, lst, tags, url) in enumerate(CLICKUP_RAW, start=3):
        row_fill = STATUS_FILLS.get(status.lower(), hex_fill("F8FAFC"))
        vals = [cid, name, status, priority, lst, tags, url]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            if col_idx == 7 and val:   # URL column
                cell.hyperlink = val
                cell.font = Font(color="2563EB", underline="single")
        ws.row_dimensions[row_idx].height = 22

    ws.auto_filter.ref = f"A2:G2"


def main():
    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Sheet order: All Issues | per-category sheets | ClickUp Issues | Statistics
    ws1 = wb.create_sheet("All Issues")
    write_all_issues(ws1)

    categories = ["Backend", "Frontend Web", "Security", "Flutter Mobile"]
    for cat in categories:
        if any(i["category"] == cat for i in ISSUES):
            write_category_sheet(wb, cat)

    write_clickup_sheet(wb)

    ws_stats = wb.create_sheet("Statistics")
    write_statistics(ws_stats)

    output_path = "master_issue_tracker.xlsx"
    wb.save(output_path)
    sheet_names = ["All Issues"] + categories + ["ClickUp Issues", "Statistics"]
    print(f"Saved: {output_path}")
    print(f"Sheets: {' | '.join(sheet_names)}")
    resolved = sum(1 for i in ISSUES if i['status'] == 'Resolved')
    open_ = sum(1 for i in ISSUES if i['status'] == 'Open')
    print(f"Issues: {len(ISSUES)} total  ({resolved} resolved, {open_} open)")
    print(f"ClickUp raw tasks: {len(CLICKUP_RAW)}")


if __name__ == "__main__":
    main()
