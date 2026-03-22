"""
generate_improvements.py — Vela Manifest Coach Feature Improvements
Run: pip install openpyxl && python generate_improvements.py
Output: tasks/vela_improvements.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

# ---------------------------------------------------------------------------
# DATA — all improvements from Manifest Coach codebase analysis
# ---------------------------------------------------------------------------

IMPROVEMENTS = [
    # ── TIER 1 — Surface Existing Data (Low effort, high impact) ─────────────
    {
        "id": "IMP-001",
        "title": "Display Study Streak Badge",
        "tier": "Tier 1 — Quick Win",
        "category": "Flutter Mobile",
        "priority": "High",
        "effort": "Low (1–2 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "UI Enhancement",
        "affected_files": (
            "vela_flutter/lib/providers/progress_provider.dart, "
            "vela_flutter/lib/ui/screens/sessions/timeline_screen.dart, "
            "vela_flutter/lib/ui/screens/profile/profile_screen.dart"
        ),
        "description": (
            "ProgressProvider already computes `streak`, `studyDays`, and `studyStreak` "
            "but these values are never displayed in any UI. Add a streak badge (e.g. '🔥 7-day streak') "
            "to the Timeline screen header and the Profile screen stats section."
        ),
        "technical_steps": (
            "1. Read progress_provider.dart — confirm streak field name. "
            "2. In timeline_screen.dart: consume ProgressProvider via ref.watch, "
            "   add VelaBadge(label: '🔥 ${state.streak} day streak') in the AppBar actions. "
            "3. In profile_screen.dart: add streak stat card alongside existing stats. "
            "4. No backend changes needed — data already fetched."
        ),
        "api_changes": "None",
        "db_changes": "None",
        "flutter_impact": "Yes — timeline_screen.dart, profile_screen.dart",
        "web_impact": "Optional — can mirror in ProfilePage.jsx",
        "psychology": "Habit loop reinforcement (James Clear). Streak visibility drives consistency (Duolingo model).",
        "new_endpoint": "No",
        "new_table": "No",
    },
    {
        "id": "IMP-002",
        "title": "Goal Progress Bar (% toward target hours)",
        "tier": "Tier 1 — Quick Win",
        "category": "Flutter Mobile + React Web",
        "priority": "High",
        "effort": "Low (1–2 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "UI Enhancement",
        "affected_files": (
            "vela_flutter/lib/ui/screens/goals/goals_screen.dart, "
            "frontend-web/src/components/GoalsPage.jsx, "
            "frontend-web/src/components/GoalsPage.css"
        ),
        "description": (
            "Goals have `target_hours` stored in DB. The goals API JOIN already returns "
            "`total_minutes` from linked study sessions. Calculate percentage = "
            "(total_minutes / (target_hours * 60)) * 100 and render a progress bar on each GoalCard."
        ),
        "technical_steps": (
            "1. Verify backend/routes/goals.js returns total_minutes in GET / response. "
            "2. Flutter: In GoalCard widget, add LinearProgressIndicator with "
            "   value = min(totalMinutes / (targetHours * 60), 1.0). "
            "3. Web: In GoalsPage.jsx GoalCard section, add <progress> or CSS bar element. "
            "4. Show percentage label: '67% complete (20h / 30h target)'."
        ),
        "api_changes": "None (data already in response — verify total_minutes field)",
        "db_changes": "None",
        "flutter_impact": "Yes — goals_screen.dart / GoalCard widget",
        "web_impact": "Yes — GoalsPage.jsx",
        "psychology": "Zeigarnik Effect — visible progress gaps keep goals motivating. Expectancy Theory — seeing progress builds belief.",
        "new_endpoint": "No",
        "new_table": "No",
    },
    {
        "id": "IMP-003",
        "title": "Task Completion Celebration Animation",
        "tier": "Tier 1 — Quick Win",
        "category": "Flutter Mobile + React Web",
        "priority": "Medium",
        "effort": "Low (1 day)",
        "impact": "Medium",
        "status": "Pending",
        "feature_type": "UI Enhancement",
        "affected_files": (
            "vela_flutter/lib/ui/screens/tasks/tasks_screen.dart, "
            "frontend-web/src/components/Tasks.jsx, "
            "frontend-web/src/components/Tasks.css"
        ),
        "description": (
            "When a task is marked complete, nothing happens visually beyond a strikethrough. "
            "Add a brief confetti burst or checkmark animation to create a dopamine reward moment. "
            "Flutter: use the `confetti` package. Web: CSS keyframe animation or canvas confetti."
        ),
        "technical_steps": (
            "1. Flutter: Add `confetti: ^0.7.0` to pubspec.yaml. "
            "   Create ConfettiController, trigger on task toggle complete. "
            "   Respect AppAnimations.reducedMotion — skip if user prefers-reduced-motion. "
            "2. Web: On task complete in Tasks.jsx, trigger a CSS animation class for 600ms. "
            "   Use @keyframes for a simple scale+fade or import canvas-confetti npm package. "
            "   Wrap in prefers-reduced-motion media query check."
        ),
        "api_changes": "None",
        "db_changes": "None",
        "flutter_impact": "Yes — tasks_screen.dart",
        "web_impact": "Yes — Tasks.jsx",
        "psychology": "Operant conditioning — immediate positive reinforcement after behavior increases repetition (B.F. Skinner).",
        "new_endpoint": "No",
        "new_table": "No",
    },

    # ── TIER 2 — Wire Existing Backend (Medium effort) ────────────────────────
    {
        "id": "IMP-004",
        "title": "Surface Journal Feature (Goal-Linked Journaling)",
        "tier": "Tier 2 — Wire Backend",
        "category": "React Web + Flutter Mobile",
        "priority": "Critical",
        "effort": "Medium (3–5 days)",
        "impact": "Very High",
        "status": "Pending",
        "feature_type": "New UI for Existing Backend",
        "affected_files": (
            "backend/routes/journal.js (READ FIRST — already complete), "
            "frontend-web/src/api.js (journalApi already exists), "
            "frontend-web/src/components/GoalsPage.jsx, "
            "frontend-web/src/components/GoalDetailModal.jsx (new Journal tab), "
            "vela_flutter/lib/ui/screens/goals/goals_screen.dart, "
            "vela_flutter/lib/data/repositories/ (new journal_repository.dart), "
            "vela_flutter/lib/providers/ (new journal_provider.dart)"
        ),
        "description": (
            "The journal backend is fully built (journal.js: GET, POST, PUT, DELETE, stats endpoint). "
            "The journalApi client in api.js is complete. Zero backend work needed. "
            "Build the UI layer: a Journal tab inside GoalDetailModal (web) and "
            "a journal section/screen inside GoalsScreen (Flutter)."
        ),
        "technical_steps": (
            "1. Read backend/routes/journal.js — note exact schema fields (entry_text, mood, type, goal_id). "
            "2. Web: In GoalDetailModal.jsx add a 'Journal' tab using the existing Tabs component. "
            "   Load entries via journalApi.getByGoal(goalId). Render entry list + add entry form. "
            "3. Web: Create JournalEntry.jsx card component (date, mood, text). "
            "4. Flutter: Create JournalRepository using DioClient (GET /api/journal/goal/:id, POST). "
            "5. Flutter: Create JournalProvider (StateNotifier pattern matching other providers). "
            "6. Flutter: Add journal entries section to goals detail view. "
            "7. Both: Wire journalApi.getStats() to show mood trend on goal card."
        ),
        "api_changes": "None — backend/routes/journal.js already mounted and complete",
        "db_changes": "None",
        "flutter_impact": "Yes — new journal_repository.dart, journal_provider.dart, journal UI",
        "web_impact": "Yes — GoalDetailModal.jsx gets Journal tab",
        "psychology": "Journaling = scripting (The Secret) + cognitive reappraisal (CBT) + self-monitoring. #1 technique across all manifestation systems.",
        "new_endpoint": "No (already exists)",
        "new_table": "No (journal_entries table already exists)",
    },
    {
        "id": "IMP-005",
        "title": "Contextualized AI Ask Coach",
        "tier": "Tier 2 — Wire Backend",
        "category": "React Web + Flutter Mobile + Backend",
        "priority": "High",
        "effort": "Medium (2–3 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "Feature Enhancement",
        "affected_files": (
            "backend/routes/ask.js, "
            "frontend-web/src/components/AskPage.jsx, "
            "frontend-web/src/api.js (askApi.chat), "
            "vela_flutter/lib/data/repositories/ask_repository.dart, "
            "vela_flutter/lib/ui/screens/ask/ask_screen.dart"
        ),
        "description": (
            "The Ask AI screen (web + Flutter) is a generic chat interface with no knowledge of the user. "
            "Inject a system context block before each chat message containing: active subject name, "
            "pending task count + top 3 task titles, current streak, recent session count (last 7 days), "
            "and active goals. This turns a generic chatbot into a personal study coach."
        ),
        "technical_steps": (
            "1. In backend/routes/ask.js: Accept optional `context` field in POST body. "
            "   Prepend context as a system message before the user message in the LLM call. "
            "2. In frontend AskPage.jsx / ask_screen.dart: Before sending a message, "
            "   fetch: currentSubject (from SubjectsProvider), activeTasks (top 3 from TasksProvider), "
            "   streak (from ProgressProvider), recentSessions count. "
            "3. Bundle as context object and pass to askApi.chat(messages, context). "
            "4. Keep context injection optional — degrade gracefully if data not loaded. "
            "5. Add manifest-coach.md system prompt content to ask.js as the AI persona."
        ),
        "api_changes": "Non-breaking — add optional `context` field to POST /api/ask body",
        "db_changes": "None",
        "flutter_impact": "Yes — ask_screen.dart, ask_repository.dart",
        "web_impact": "Yes — AskPage.jsx",
        "psychology": "A coach who knows your situation can coach you. Contextual AI = RAS priming + personalized feedback loop.",
        "new_endpoint": "No (enhance existing /api/ask)",
        "new_table": "No",
    },
    {
        "id": "IMP-006",
        "title": "Flutter Goal Detail Modal (Parity Gap)",
        "tier": "Tier 2 — Wire Backend",
        "category": "Flutter Mobile",
        "priority": "High",
        "effort": "Medium (2–3 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "Parity Gap Fix",
        "affected_files": (
            "vela_flutter/lib/ui/screens/goals/goals_screen.dart, "
            "vela_flutter/lib/ui/screens/goals/goal_detail_modal.dart (new file), "
            "vela_flutter/lib/data/repositories/goals_repository.dart"
        ),
        "description": (
            "Web has GoalDetailModal.jsx (full detail view, linked sessions, journal). "
            "Flutter tapping a goal does nothing — no detail view. "
            "Create GoalDetailModal as a bottom sheet matching web functionality: "
            "goal title, description, status, target date, progress bar, linked sessions list."
        ),
        "technical_steps": (
            "1. Read frontend-web/src/components/GoalDetailModal.jsx for web spec. "
            "2. Create goal_detail_modal.dart as a VelaModal bottom sheet. "
            "3. Sections: Header (title, status badge, close button), "
            "   Progress bar (hours_logged / target_hours), "
            "   Target date chip, Description text, "
            "   Linked sessions list (load via GET /api/progress/sessions?goal_id=X). "
            "4. Wire from GoalsScreen: onTap → showModalBottomSheet(GoalDetailModal). "
            "5. Add Edit + Delete actions in modal footer."
        ),
        "api_changes": "None",
        "db_changes": "None",
        "flutter_impact": "Yes — goals_screen.dart + new goal_detail_modal.dart",
        "web_impact": "None",
        "psychology": "Goals without detail view feel abandoned. Long-term orientation requires regular goal review (Expectancy Theory).",
        "new_endpoint": "No",
        "new_table": "No",
    },

    # ── TIER 3 — New Manifestation Layer ─────────────────────────────────────
    {
        "id": "IMP-007",
        "title": "Morning Intention Modal (Daily Focus Setter)",
        "tier": "Tier 3 — New Feature",
        "category": "Flutter Mobile + React Web",
        "priority": "High",
        "effort": "Medium (3–4 days)",
        "impact": "Very High",
        "status": "Pending",
        "feature_type": "New Feature",
        "affected_files": (
            "backend/routes/journal.js (add type='intention' support — check schema), "
            "vela_flutter/lib/ui/screens/shared/app_shell.dart, "
            "vela_flutter/lib/core/storage/local_storage.dart (store last_intention_date), "
            "frontend-web/src/App.jsx (trigger on first daily load), "
            "frontend-web/src/components/MorningIntentionModal.jsx (new), "
            "frontend-web/src/components/MorningIntentionModal.css (new)"
        ),
        "description": (
            "On first app open of the day, show a fullscreen modal: "
            "'What are you manifesting today?' with a text input and optional goal selector. "
            "Save the intention as a journal_entry with type='intention'. "
            "Dismiss until tomorrow (store last_shown_date in localStorage / SharedPreferences). "
            "Show the intention as a banner at top of Tasks screen for the rest of the day."
        ),
        "technical_steps": (
            "1. Check backend/routes/journal.js schema — does journal_entries have a `type` column? "
            "   If not: ALTER TABLE journal_entries ADD COLUMN type VARCHAR(50) DEFAULT 'reflection'; "
            "2. Flutter: In app_shell.dart initState, check local_storage for last_intention_date. "
            "   If not today, show IntentionModal as overlay after auth check. "
            "3. Flutter: Create IntentionModal widget — title, text field, goal dropdown, Save button. "
            "   On save: POST /api/journal with type='intention', store today's date in local_storage. "
            "4. Web: In App.jsx useEffect (on mount), check localStorage for last_intention_date. "
            "   If not today, render MorningIntentionModal. "
            "5. Both: After save, show intention as a dismissable banner in Tasks screen."
        ),
        "api_changes": "Minor — POST /api/journal gains optional type field (non-breaking)",
        "db_changes": "Add `type VARCHAR(50) DEFAULT 'reflection'` to journal_entries (nullable safe)",
        "flutter_impact": "Yes — app_shell.dart, new intention_modal.dart",
        "web_impact": "Yes — App.jsx, new MorningIntentionModal.jsx",
        "psychology": "Implementation intentions (Peter Gollwitzer): 'When X, I will do Y' increases follow-through by 2x. RAS priming for the day.",
        "new_endpoint": "No (POST /api/journal already exists)",
        "new_table": "No (ALTER existing journal_entries)",
    },
    {
        "id": "IMP-008",
        "title": "Post-Session Reflection Prompt",
        "tier": "Tier 3 — New Feature",
        "category": "Flutter Mobile + React Web",
        "priority": "Medium",
        "effort": "Low–Medium (2 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "New Feature",
        "affected_files": (
            "vela_flutter/lib/ui/screens/sessions/add_session_modal.dart, "
            "frontend-web/src/components/AddSessionModal.jsx (or wherever session is saved)"
        ),
        "description": (
            "After a study session is successfully logged, show a brief prompt: "
            "'Great session! Want to note what you manifested?' with a 2-tap flow: "
            "Skip | Add reflection. If Add: show a 1-field text input, "
            "save as journal_entry linked to the session (goal_id if session has one)."
        ),
        "technical_steps": (
            "1. Flutter: In add_session_modal.dart, on successful POST /api/progress/sessions, "
            "   show a SnackBar/dialog: 'Session logged! Add a reflection?' with action button. "
            "2. If tapped: show a simple bottom sheet with text field + Save. "
            "   POST /api/journal with entry_text, type='reflection', goal_id (from session). "
            "3. Web: After session save success in App.jsx or Timeline.jsx, "
            "   show a lightweight modal with same prompt. "
            "4. Both: Reflection is optional — 'Skip' dismisses with no action."
        ),
        "api_changes": "None",
        "db_changes": "None (uses existing journal_entries table)",
        "flutter_impact": "Yes — add_session_modal.dart",
        "web_impact": "Yes — wherever AddSessionModal success is handled",
        "psychology": "Spaced retrieval + emotional tagging. Reflecting on sessions consolidates learning memory (Cognitive psychology). Evening review habit.",
        "new_endpoint": "No",
        "new_table": "No",
    },
    {
        "id": "IMP-009",
        "title": "Daily Affirmation / Motivational Banner",
        "tier": "Tier 3 — New Feature",
        "category": "Flutter Mobile + React Web",
        "priority": "Medium",
        "effort": "Low (1 day)",
        "impact": "Medium",
        "status": "Pending",
        "feature_type": "New Feature",
        "affected_files": (
            "vela_flutter/lib/ui/screens/tasks/tasks_screen.dart, "
            "frontend-web/src/components/Tasks.jsx, "
            "vela_flutter/lib/core/utils/ (new affirmations.dart), "
            "frontend-web/src/utils/ (new affirmations.js)"
        ),
        "description": (
            "Add a rotating daily affirmation/motivational quote at the top of the Tasks screen. "
            "Curated list of 30+ study/manifestation affirmations. "
            "Rotate daily using day-of-year % list.length (deterministic, same quote all day). "
            "Dismissable per-day. No backend needed — pure frontend."
        ),
        "technical_steps": (
            "1. Create affirmations.dart / affirmations.js with curated list of 30 affirmations. "
            "2. Deterministic selection: dayOfYear % affirmations.length. "
            "3. Flutter: Add a VelaCard with italic quote text at top of TasksScreen ListView. "
            "   Store dismissed date in local_storage — hide for rest of day if dismissed. "
            "4. Web: In Tasks.jsx render a dismissable banner above task list. "
            "   Use localStorage('affirmation_dismissed_date') to persist dismissal."
        ),
        "api_changes": "None",
        "db_changes": "None",
        "flutter_impact": "Yes — tasks_screen.dart",
        "web_impact": "Yes — Tasks.jsx",
        "psychology": "Positive self-talk primes identity-based behavior (James Clear). Daily repetition builds belief schemas (CBT affirmations).",
        "new_endpoint": "No",
        "new_table": "No",
    },
    {
        "id": "IMP-010",
        "title": "Flutter Subtask Toggle + Completion (Parity Gap)",
        "tier": "Tier 2 — Wire Backend",
        "category": "Flutter Mobile",
        "priority": "High",
        "effort": "Low–Medium (1–2 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "Parity Gap Fix",
        "affected_files": (
            "vela_flutter/lib/ui/screens/tasks/task_detail_modal.dart, "
            "vela_flutter/lib/providers/tasks_provider.dart, "
            "vela_flutter/lib/data/repositories/tasks_repository.dart"
        ),
        "description": (
            "Subtasks in Flutter TaskDetailModal are currently read-only (shown but not tappable). "
            "Web TaskDetailModal has full toggle complete + delete on each subtask. "
            "Wire subtask checkboxes to PUT /api/tasks/:id with completed=true/false."
        ),
        "technical_steps": (
            "1. Read backend/routes/tasks.js PUT /:id endpoint — confirm it accepts completed field. "
            "2. In task_detail_modal.dart: Wrap each subtask row in GestureDetector/InkWell. "
            "3. On tap: call tasksProvider.toggleTaskComplete(subtaskId) — reuse existing method. "
            "4. Update local state to reflect toggle immediately (optimistic UI). "
            "5. Add delete icon on subtask row → call tasksProvider.deleteTask(subtaskId)."
        ),
        "api_changes": "None (PUT /api/tasks/:id already handles completed field)",
        "db_changes": "None",
        "flutter_impact": "Yes — task_detail_modal.dart, tasks_provider.dart",
        "web_impact": "None",
        "psychology": "Completing subtasks = micro-wins. Momentum builds on small victories (Progress Principle — Teresa Amabile).",
        "new_endpoint": "No",
        "new_table": "No",
    },
    {
        "id": "IMP-011",
        "title": "Sessions Progress Graph in Flutter (Parity Gap)",
        "tier": "Tier 2 — Wire Backend",
        "category": "Flutter Mobile",
        "priority": "Medium",
        "effort": "Medium (2–3 days)",
        "impact": "Medium",
        "status": "Pending",
        "feature_type": "Parity Gap Fix",
        "affected_files": (
            "vela_flutter/lib/ui/screens/sessions/timeline_screen.dart, "
            "vela_flutter/lib/ui/screens/sessions/sessions_graph_screen.dart (new), "
            "vela_flutter/lib/providers/progress_provider.dart"
        ),
        "description": (
            "Web has SessionsGraphModal showing study time trends over time. "
            "Flutter Timeline screen has no graph. "
            "Add a chart icon button to Timeline AppBar that opens a graph screen "
            "showing daily/weekly study minutes using fl_chart package."
        ),
        "technical_steps": (
            "1. Add `fl_chart: ^0.68.0` to pubspec.yaml. "
            "2. Create sessions_graph_screen.dart with BarChart or LineChart. "
            "   Data source: ProgressProvider sessions grouped by date. "
            "3. X-axis: last 14 days. Y-axis: minutes studied per day. "
            "4. Add IconButton(icon: Icon(Icons.bar_chart)) to timeline_screen.dart AppBar. "
            "   On tap: navigate to SessionsGraphScreen."
        ),
        "api_changes": "None",
        "db_changes": "None",
        "flutter_impact": "Yes — timeline_screen.dart + new sessions_graph_screen.dart",
        "web_impact": "None",
        "psychology": "Progress visualization activates Expectancy Theory — seeing results builds belief in effort→outcome causation.",
        "new_endpoint": "No",
        "new_table": "No",
    },
    {
        "id": "IMP-012",
        "title": "Flutter Integration Test Suite",
        "tier": "Tier 2 — Wire Backend",
        "category": "Flutter Mobile",
        "priority": "High",
        "effort": "Medium (3–5 days)",
        "impact": "High",
        "status": "Pending",
        "feature_type": "Testing Infrastructure",
        "affected_files": (
            "vela_flutter/integration_test/app_test.dart (new), "
            "vela_flutter/pubspec.yaml, "
            "vela_flutter/lib/main.dart, "
            ".github/workflows/deploy.yml"
        ),
        "description": (
            "Set up the `integration_test` package to run end-to-end UI tests on a real device or emulator. "
            "Tests run against the real backend API (not mocked). "
            "Cover 6 critical user flows: login, add task, log study session, switch subject, "
            "add attachment, and create note. "
            "Add Key() identifiers to critical widgets to make them findable in tests."
        ),
        "technical_steps": (
            "1. Add `integration_test` and `flutter_test` to pubspec.yaml dev_dependencies. "
            "2. Create integration_test/ directory with app_test.dart as the entry point. "
            "3. Add Key() identifiers to: login form fields (email, password), "
            "   bottom nav tabs, FABs, and modal action buttons across all screens. "
            "4. Write testWidgets for: "
            "   - Login flow: enter credentials → reach home screen. "
            "   - Task creation: tap FAB → fill modal → task appears in list. "
            "   - Session log: fill session form → appears in timeline. "
            "   - Subject switch: switch subject → tasks/sessions update. "
            "   - Attachment upload: pick file → appears in attachments tab. "
            "   - Note creation: create note → appears in notes list. "
            "5. Add `flutter test integration_test/` step to GitHub Actions deploy.yml. "
            "6. Document run command in vela_flutter/README.md."
        ),
        "api_changes": "None — tests consume existing API endpoints",
        "db_changes": "None",
        "flutter_impact": "Yes — integration_test/, main.dart, all screen widgets (Key additions)",
        "web_impact": "No",
        "psychology": "N/A — Engineering quality gate. Prevents regressions across critical user flows.",
        "new_endpoint": "No",
        "new_table": "No",
    },
]

# ---------------------------------------------------------------------------
# COLOR PALETTE
# ---------------------------------------------------------------------------
COLORS = {
    "header_bg":       "1E293B",  # Dark slate
    "header_fg":       "FFFFFF",
    "tier1_bg":        "DCFCE7",  # Light green
    "tier1_accent":    "16A34A",  # Green
    "tier2_bg":        "DBEAFE",  # Light blue
    "tier2_accent":    "2563EB",  # Blue
    "tier3_bg":        "FEF9C3",  # Light yellow
    "tier3_accent":    "CA8A04",  # Yellow
    "critical_bg":     "FEE2E2",  # Light red
    "critical_fg":     "DC2626",
    "high_bg":         "FEF3C7",  # Light orange
    "high_fg":         "D97706",
    "medium_bg":       "E0F2FE",  # Light sky
    "medium_fg":       "0284C7",
    "alt_row":         "F8FAFC",  # Off-white
    "border":          "CBD5E1",
    "section_header":  "334155",
    "summary_bg":      "0F172A",
}

TIER_COLORS = {
    "Tier 1 — Quick Win":     ("DCFCE7", "166534"),
    "Tier 2 — Wire Backend":  ("DBEAFE", "1E40AF"),
    "Tier 3 — New Feature":   ("FEF9C3", "713F12"),
}

PRIORITY_COLORS = {
    "Critical": ("FEE2E2", "991B1B"),
    "High":     ("FEF3C7", "92400E"),
    "Medium":   ("E0F2FE", "075985"),
    "Low":      ("F0FDF4", "166534"),
}

IMPACT_COLORS = {
    "Very High": ("FAE8FF", "7E22CE"),
    "High":      ("DCFCE7", "166534"),
    "Medium":    ("E0F2FE", "075985"),
    "Low":       ("F1F5F9", "475569"),
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def make_border(style="thin"):
    s = Side(style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value, bg="1E293B", fg="FFFFFF", size=11, bold=True, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = make_border()
    return cell

def data_cell(ws, row, col, value, bg="FFFFFF", bold=False, wrap=True, align="left", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color="1E293B", size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = make_border()
    return cell

def color_cell(ws, row, col, value, bg_fg_tuple, bold=True):
    bg, fg = bg_fg_tuple
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=10, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = make_border()
    return cell

# ---------------------------------------------------------------------------
# SHEET 1: ALL IMPROVEMENTS
# ---------------------------------------------------------------------------
def build_main_sheet(wb):
    ws = wb.create_sheet("All Improvements")
    ws.sheet_view.showGridLines = False

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    title = ws.cell(row=1, column=1,
        value="VELA — Manifest Coach Feature Improvements Roadmap")
    title.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
    title.fill = PatternFill("solid", fgColor=COLORS["summary_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:N2")
    sub = ws.cell(row=2, column=1,
        value=f"Generated from codebase analysis · {len(IMPROVEMENTS)} improvements · Tiers 1–3")
    sub.font = Font(italic=True, color="94A3B8", size=10, name="Calibri")
    sub.fill = PatternFill("solid", fgColor=COLORS["summary_bg"])
    sub.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20

    # ── Column headers ───────────────────────────────────────────────────────
    COLS = [
        ("ID", 10),
        ("Title", 38),
        ("Tier", 22),
        ("Category", 24),
        ("Priority", 12),
        ("Effort", 18),
        ("Impact", 12),
        ("New Endpoint?", 14),
        ("New DB Table?", 14),
        ("API Changes", 26),
        ("DB Changes", 26),
        ("Flutter Impact", 28),
        ("Web Impact", 28),
        ("Status", 14),
    ]
    for col_idx, (label, width) in enumerate(COLS, 1):
        header_cell(ws, 3, col_idx, label)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    for r_idx, imp in enumerate(IMPROVEMENTS):
        row = r_idx + 4
        bg = TIER_COLORS.get(imp["tier"], ("FFFFFF", "000000"))[0]
        alt = bg  # use tier color for row tinting

        data_cell(ws, row, 1,  imp["id"],           bg=alt, bold=True, wrap=False, align="center")
        data_cell(ws, row, 2,  imp["title"],         bg="FFFFFF", bold=True)
        color_cell(ws, row, 3, imp["tier"],          TIER_COLORS.get(imp["tier"], ("F8FAFC","334155")))
        data_cell(ws, row, 4,  imp["category"],      bg=alt)
        color_cell(ws, row, 5, imp["priority"],      PRIORITY_COLORS.get(imp["priority"], ("F8FAFC","334155")))
        data_cell(ws, row, 6,  imp["effort"],        bg=alt, align="center")
        color_cell(ws, row, 7, imp["impact"],        IMPACT_COLORS.get(imp["impact"], ("F8FAFC","334155")))
        data_cell(ws, row, 8,  imp["new_endpoint"],  bg=alt, align="center")
        data_cell(ws, row, 9,  imp["new_table"],     bg=alt, align="center")
        data_cell(ws, row, 10, imp["api_changes"],   bg="FFFFFF")
        data_cell(ws, row, 11, imp["db_changes"],    bg="FFFFFF")
        data_cell(ws, row, 12, imp["flutter_impact"],bg=alt)
        data_cell(ws, row, 13, imp["web_impact"],    bg=alt)
        color_cell(ws, row, 14, imp["status"],       ("FEF9C3","92400E") if imp["status"]=="Pending" else ("DCFCE7","166534"))

        ws.row_dimensions[row].height = 30

    ws.freeze_panes = "A4"
    return ws

# ---------------------------------------------------------------------------
# SHEET 2: TECHNICAL DETAIL
# ---------------------------------------------------------------------------
def build_detail_sheet(wb):
    ws = wb.create_sheet("Technical Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="VELA — Technical Implementation Detail")
    t.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    t.fill = PatternFill("solid", fgColor=COLORS["summary_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    COLS = [
        ("ID",               10),
        ("Title",            38),
        ("Description",      60),
        ("Technical Steps",  80),
        ("Psychology / Why", 55),
    ]
    for col_idx, (label, width) in enumerate(COLS, 1):
        header_cell(ws, 2, col_idx, label)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28

    for r_idx, imp in enumerate(IMPROVEMENTS):
        row = r_idx + 3
        bg = TIER_COLORS.get(imp["tier"], ("FFFFFF", "000000"))[0]

        data_cell(ws, row, 1, imp["id"],               bg=bg, bold=True, align="center")
        data_cell(ws, row, 2, imp["title"],             bg="FFFFFF", bold=True)
        data_cell(ws, row, 3, imp["description"],       bg="FAFAFA")
        data_cell(ws, row, 4, imp["technical_steps"],   bg="FAFAFA")
        data_cell(ws, row, 5, imp["psychology"],        bg=bg)

        ws.row_dimensions[row].height = 90

    ws.freeze_panes = "A3"
    return ws

# ---------------------------------------------------------------------------
# SHEET 3: FILES AFFECTED
# ---------------------------------------------------------------------------
def build_files_sheet(wb):
    ws = wb.create_sheet("Files Affected")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="VELA — Files to be Modified per Improvement")
    t.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    t.fill = PatternFill("solid", fgColor=COLORS["summary_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    COLS = [("ID", 12), ("Title", 38), ("Files Affected", 90)]
    for col_idx, (label, width) in enumerate(COLS, 1):
        header_cell(ws, 2, col_idx, label)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28

    for r_idx, imp in enumerate(IMPROVEMENTS):
        row = r_idx + 3
        bg = TIER_COLORS.get(imp["tier"], ("FFFFFF", "000000"))[0]
        data_cell(ws, row, 1, imp["id"],              bg=bg, bold=True, align="center")
        data_cell(ws, row, 2, imp["title"],           bg="FFFFFF", bold=True)
        data_cell(ws, row, 3, imp["affected_files"],  bg="FAFAFA")
        ws.row_dimensions[row].height = 60

    ws.freeze_panes = "A3"
    return ws

# ---------------------------------------------------------------------------
# SHEET 4: SUMMARY DASHBOARD
# ---------------------------------------------------------------------------
def build_summary_sheet(wb):
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    def title_row(row, text, bg="1E293B"):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 26

    def stat_row(row, label, value, bg="F8FAFC", value_bg="FFFFFF"):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(color="475569", size=10, name="Calibri")
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.border = make_border()
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"B{row}:D{row}")
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = Font(bold=True, color="1E293B", size=11, name="Calibri")
        c2.fill = PatternFill("solid", fgColor=value_bg)
        c2.border = make_border()
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # Title
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="VELA Improvements — Summary Dashboard")
    t.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
    t.fill = PatternFill("solid", fgColor=COLORS["summary_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Counts
    title_row(3, "  Improvement Counts by Tier", "1E3A5F")
    tiers = {}
    for imp in IMPROVEMENTS:
        tiers[imp["tier"]] = tiers.get(imp["tier"], 0) + 1
    r = 4
    for tier, count in tiers.items():
        bg, fg = TIER_COLORS.get(tier, ("F8FAFC","334155"))
        stat_row(r, tier, f"{count} improvements", bg=bg, value_bg=bg)
        r += 1

    title_row(r + 1, "  Priority Breakdown", "1E3A5F")
    r += 2
    priorities = {}
    for imp in IMPROVEMENTS:
        priorities[imp["priority"]] = priorities.get(imp["priority"], 0) + 1
    for pri, count in sorted(priorities.items()):
        bg, fg = PRIORITY_COLORS.get(pri, ("F8FAFC","334155"))
        stat_row(r, pri, f"{count} improvements", bg=bg, value_bg=bg)
        r += 1

    title_row(r + 1, "  Backend / DB Impact Summary", "1E3A5F")
    r += 2
    no_endpoint = sum(1 for i in IMPROVEMENTS if i["new_endpoint"] == "No")
    no_table = sum(1 for i in IMPROVEMENTS if i["new_table"] == "No")
    stat_row(r,     "Improvements needing NO new endpoint", f"{no_endpoint} / {len(IMPROVEMENTS)}", value_bg="DCFCE7")
    stat_row(r + 1, "Improvements needing NO new DB table",  f"{no_table} / {len(IMPROVEMENTS)}",   value_bg="DCFCE7")
    stat_row(r + 2, "Total improvements",                    str(len(IMPROVEMENTS)),                 value_bg="E0F2FE")

    return ws

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_summary_sheet(wb)
    build_main_sheet(wb)
    build_detail_sheet(wb)
    build_files_sheet(wb)

    # Set Summary as active
    wb.active = wb["Summary Dashboard"]

    output_path = "vela_improvements.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Improvements: {len(IMPROVEMENTS)}")

if __name__ == "__main__":
    main()
